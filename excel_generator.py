"""
Módulo para geração de planilhas Excel com análise de currículos
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any
import logging
from datetime import datetime

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Classe para geração de planilhas Excel com análise de currículos"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_analysis_report(self, 
                             candidates_data: List[Dict[str, Any]], 
                             job_profile: Dict[str, List[str]], 
                             output_file: str = None) -> str:
        """
        Cria relatório de análise em Excel
        
        Args:
            candidates_data: Lista com dados dos candidatos
            job_profile: Perfil da vaga
            output_file: Nome do arquivo de saída (opcional)
            
        Returns:
            Caminho do arquivo gerado
        """
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"analise_curriculos_{timestamp}.xlsx"
        
        # Criar DataFrame
        df = self._create_dataframe(candidates_data, job_profile)
        
        # Criar workbook
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Análise de Currículos"
        
        # Adicionar dados ao worksheet
        self._add_data_to_worksheet(df)
        
        # Aplicar formatação
        self._apply_formatting(df, job_profile)
        
        # Salvar arquivo
        self.workbook.save(output_file)
        logger.info(f"Relatório salvo em: {output_file}")
        
        return output_file
    
    def _create_dataframe(self, candidates_data: List[Dict[str, Any]], job_profile: Dict[str, List[str]]) -> pd.DataFrame:
        """
        Cria DataFrame com dados dos candidatos
        
        Args:
            candidates_data: Lista com dados dos candidatos
            job_profile: Perfil da vaga
            
        Returns:
            DataFrame com dados formatados
        """
        rows = []
        
        for candidate in candidates_data:
            row = {
                'Nome': candidate.get('contato', {}).get('nome', 'Não informado'),
                'E-mail': candidate.get('contato', {}).get('email', 'Não informado'),
                'Telefone': candidate.get('contato', {}).get('telefone', 'Não informado'),
                'Arquivo': candidate.get('arquivo', ''),
                'Pontuação Total': candidate.get('pontuacao_total', 0)
            }
            
            # Adicionar pontuações dos atributos
            pontuacoes = candidate.get('analise', {}).get('pontuacoes', {})
            for attr in job_profile['requeridos'] + job_profile['desejaveis']:
                row[f'Nota - {attr}'] = pontuacoes.get(attr, 0)
            
            # Adicionar resumo
            row['Resumo das Qualidades'] = candidate.get('analise', {}).get('resumo', '')
            
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # Ordenar por pontuação total (maior para menor)
        df = df.sort_values('Pontuação Total', ascending=False)
        
        return df
    
    def _add_data_to_worksheet(self, df: pd.DataFrame):
        """
        Adiciona dados do DataFrame ao worksheet
        
        Args:
            df: DataFrame com dados
        """
        # Adicionar cabeçalho
        header_row = 1
        for col_num, column_name in enumerate(df.columns, 1):
            cell = self.worksheet.cell(row=header_row, column=col_num, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adicionar dados
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Formatação especial para pontuações
                if 'Nota -' in str(self.worksheet.cell(row=1, column=col_num).value):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Colorir células baseado na pontuação
                    if isinstance(value, (int, float)) and value > 0:
                        if value >= 4:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value >= 3:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def _apply_formatting(self, df: pd.DataFrame, job_profile: Dict[str, List[str]]):
        """
        Aplica formatação ao worksheet
        
        Args:
            df: DataFrame com dados
            job_profile: Perfil da vaga
        """
        # Ajustar largura das colunas
        column_widths = {
            'A': 25,  # Nome
            'B': 30,  # E-mail
            'C': 20,  # Telefone
            'D': 30,  # Arquivo
            'E': 15,  # Pontuação Total
        }
        
        # Ajustar largura das colunas de notas
        note_col_start = 6
        for i, attr in enumerate(job_profile['requeridos'] + job_profile['desejaveis']):
            col_letter = chr(ord('A') + note_col_start + i)
            column_widths[col_letter] = 20
        
        # Ajustar largura da coluna de resumo
        resumo_col = chr(ord('A') + len(df.columns) - 1)
        column_widths[resumo_col] = 50
        
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width
        
        # Adicionar bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Congelar primeira linha
        self.worksheet.freeze_panes = 'A2'
        
        # Adicionar filtros
        self.worksheet.auto_filter.ref = f"A1:{chr(ord('A') + len(df.columns) - 1)}{len(df) + 1}"
    
    def create_summary_sheet(self, candidates_data: List[Dict[str, Any]], job_profile: Dict[str, List[str]]):
        """
        Cria planilha de resumo com estatísticas
        
        Args:
            candidates_data: Lista com dados dos candidatos
            job_profile: Perfil da vaga
        """
        if not self.workbook:
            return
        
        # Criar nova planilha
        summary_sheet = self.workbook.create_sheet("Resumo")
        
        # Estatísticas gerais
        total_candidates = len(candidates_data)
        avg_score = sum(c.get('pontuacao_total', 0) for c in candidates_data) / total_candidates if total_candidates > 0 else 0
        
        # Dados do resumo
        summary_data = [
            ["ESTATÍSTICAS GERAIS", ""],
            ["Total de Candidatos", total_candidates],
            ["Pontuação Média", round(avg_score, 2)],
            ["", ""],
            ["TOP 5 CANDIDATOS", ""],
        ]
        
        # Top 5 candidatos
        sorted_candidates = sorted(candidates_data, key=lambda x: x.get('pontuacao_total', 0), reverse=True)
        for i, candidate in enumerate(sorted_candidates[:5], 1):
            nome = candidate.get('contato', {}).get('nome', 'Não informado')
            score = candidate.get('pontuacao_total', 0)
            summary_data.append([f"{i}º lugar", f"{nome} - {score} pontos"])
        
        summary_data.extend([
            ["", ""],
            ["ATRIBUTOS MAIS BEM AVALIADOS", ""],
        ])
        
        # Análise por atributo
        attribute_scores = {}
        for candidate in candidates_data:
            pontuacoes = candidate.get('analise', {}).get('pontuacoes', {})
            for attr, score in pontuacoes.items():
                if attr not in attribute_scores:
                    attribute_scores[attr] = []
                attribute_scores[attr].append(score)
        
        # Calcular média por atributo
        for attr, scores in attribute_scores.items():
            avg_score = sum(scores) / len(scores) if scores else 0
            summary_data.append([attr, f"{round(avg_score, 2)} pontos"])
        
        # Adicionar dados à planilha
        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_sheet.cell(row=row_num, column=1, value=label)
            summary_sheet.cell(row=row_num, column=2, value=value)
            
            # Formatação do cabeçalho
            if row_num == 1 or row_num == 5 or row_num == len(summary_data) - len(attribute_scores):
                summary_sheet.cell(row=row_num, column=1).font = Font(bold=True)
                summary_sheet.cell(row=row_num, column=2).font = Font(bold=True)
        
        # Ajustar largura das colunas
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 30
